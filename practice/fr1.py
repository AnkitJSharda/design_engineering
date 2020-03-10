import face_recognition
import cv2
from openpyxl import Workbook
wb=Workbook()
filepath="/Users/AnkitKJS/Desktop/DESIGN-ENGINEERING/practice/demo.xlsx"

# Get a reference to webcam #0 (the default one)
video_capture = cv2.VideoCapture(0)

# Load a sample picture and learn how to recognize it.

obama_image = face_recognition.load_image_file("/Users/AnkitKJS/Desktop/DESIGN-ENGINEERING/practice/known_people/obama/obama.jpg")
obama_face_encoding = face_recognition.face_encodings(obama_image)[0]


# Load a second sample picture and learn how to recognize it.

biden_image = face_recognition.load_image_file("/Users/AnkitKJS/Desktop/DESIGN-ENGINEERING/practice/known_people/biden/biden.jpg")
biden_face_encoding = face_recognition.face_encodings(biden_image)[0]
vaibhav_image = face_recognition.load_image_file("/Users/AnkitKJS/Desktop/DESIGN-ENGINEERING/practice/known_people/vaibhav/vaibhav.jpg")
vaibhav_face_encoding = face_recognition.face_encodings(vaibhav_image)[0]
hely_image = face_recognition.load_image_file("/Users/AnkitKJS/Desktop/DESIGN-ENGINEERING/practice/known_people/hely/hely.jpg")
hely_face_encoding = face_recognition.face_encodings(hely_image)[0]
smit_image = face_recognition.load_image_file("/Users/AnkitKJS/Desktop/DESIGN-ENGINEERING/practice/known_people/smit/smit.jpg")
smit_face_encoding = face_recognition.face_encodings(smit_image)[0]
ankit_image = face_recognition.load_image_file("/Users/AnkitKJS/Desktop/DESIGN-ENGINEERING/practice/known_people/ankit/ankit.jpg")
ankit_face_encoding = face_recognition.face_encodings(ankit_image)[0]


# Create arrays of known face encodings and their names
known_face_encodings = [
    obama_face_encoding,
    biden_face_encoding,
    smit_face_encoding,
    hely_face_encoding,
    vaibhav_face_encoding,
    ankit_face_encoding
]
known_face_names = [
    "Barack Obama",
    "Joe Biden",
    "Smit Chavda",
    "Hely Parmar",
    "Vaibhav Soni",
    "Ankit Sharda"
]

# Initialize some variables
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True


sheet=wb.active
sheet.cell(row=1, column=1).value = "Biden"
sheet.cell(row=1, column=2).value = "Smit"
sheet.cell(row=1, column=3).value = "Hely"
sheet.cell(row=1, column=4).value = "Vaibhav"
sheet.cell(row=1, column=5).value = "Ankit"


while True:
    process_this_frame = True
    # Grab a single frame of video
    ret, frame = video_capture.read()

    # Resize frame of video to 1/4 size for faster face recognition processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

    # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
    rgb_small_frame = small_frame[:, :, ::-1]

    # Only process every other frame of video to save time
    if process_this_frame:
        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        face_names = []
        for face_encoding in face_encodings:
            # See if the face is a match for the known face(s)
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"

            # If a match was found in known_face_encodings, just use the first one.
            if True in matches:
                first_match_index = matches.index(True)
                #v1 =  'A' + first_match_index
                #sheet['A1'] = first_match_index
                name = known_face_names[first_match_index]
                #sheet['A1'] = first_match_index
                #sheet['v1'] = first_match_index
                sheet.cell(row=2, column=first_match_index).value = "P"
            face_names.append(name)

    process_this_frame = not process_this_frame


    # Display the results
    for (top, right, bottom, left), name in zip(face_locations, face_names):
        # Scale back up face locations since the frame we detected in was scaled to 1/4 size
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4

        # Draw a box around the face
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

        # Draw a label with a name below the face
        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

    # Display the resulting image
    cv2.imshow('Video', frame)

    # Hit 'q' on the keyboard to quit!
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release handle to the webcam
video_capture.release()
cv2.destroyAllWindows()

wb.save(filepath)

